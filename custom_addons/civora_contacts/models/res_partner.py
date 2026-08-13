# -*- coding: utf-8 -*-
from odoo import api, fields, models
from odoo.tools.translate import _
from odoo.exceptions import ValidationError


# Valeurs de consentement RGPD (alignees sur le front CIVORA).
CIVORA_CONSENT = [
    ('opt_in', "Opt-in"),
    ('opt_out', "Opt-out"),
    ('none', "—"),
]


class ResPartner(models.Model):
    """Extension CIVORA de res.partner.

    On etend le contact standard Odoo (socle invisible : nom, email, telephone,
    ville, adresse, societe, etiquettes, notes, comptabilite...) avec la couche
    metier CIVORA, alignee sur le modele du front (StoredContact) : role(s),
    source, score IA, statut, agent, prochaine action, budget, RGPD.

    Aucune vue Odoo : ces champs sont pilotes par les ecrans OWL de CIVORA.
    Correspondance avec les champs natifs reutilises :
      name / email / phone / city / street  -> natifs res.partner
      company                                -> company_name (natif)
      notes                                  -> comment (natif)
      tags                                   -> category_id (natif)
    """
    _inherit = 'res.partner'

    # --- Marqueur ------------------------------------------------------
    civora_is_contact = fields.Boolean(
        string="Contact CIVORA",
        default=False,
        index=True,
        help="Coche pour les contacts geres via l'application CIVORA.",
    )

    # --- Roles (multi) + role principal --------------------------------
    civora_role_ids = fields.Many2many(
        'civora.contact.role',
        'civora_partner_role_rel', 'partner_id', 'role_id',
        string="Roles",
        help="Un contact peut cumuler plusieurs roles (proprietaire, locataire...).",
    )
    civora_primary_role_id = fields.Many2one(
        'civora.contact.role',
        string="Role principal",
        ondelete='restrict',
        help="Role mis en avant pour l'affichage et le tri. Doit faire partie des roles.",
    )
    civora_role_names = fields.Char(
        string="Roles (libelle)",
        compute='_compute_civora_role_names',
        help="Liste des roles en texte, pour l'affichage.",
    )

    # --- Acquisition & segmentation ------------------------------------
    civora_source_id = fields.Many2one(
        'civora.contact.source',
        string="Source",
        ondelete='restrict',
        index=True,
        help="Canal d'acquisition du contact.",
    )
    civora_segment_ids = fields.Many2many(
        'civora.contact.segment',
        'civora_partner_segment_rel', 'partner_id', 'segment_id',
        string="Segments IA",
    )

    # --- Suivi commercial ----------------------------------------------
    civora_status = fields.Selection(
        [
            ('chaud', "Chaud"),
            ('actif', "Actif"),
            ('qualifie', "Qualifie"),
            ('a_risque', "A risque"),
            ('inactif', "Inactif"),
        ],
        string="Statut",
        index=True,
        help="Statut du contact dans le cycle CRM.",
    )
    civora_ai_score = fields.Integer(
        string="Score IA",
        default=0,
        help="Score d'engagement / valeur du contact, de 0 a 100.",
    )
    civora_ai_score_manual = fields.Boolean(
        string="Score verrouille (manuel)",
        default=False,
        help="Si coche, le score IA ne sera pas recalcule automatiquement par le cron.",
    )
    civora_ai_score_breakdown = fields.Text(
        string="Composition du score",
        help="Detail JSON de la derniere ventilation du score (5 composants).",
    )
    civora_ai_score_updated = fields.Datetime(
        string="Score recalcule le",
        readonly=True,
    )
    civora_agent_id = fields.Many2one(
        'res.users',
        string="Agent responsable",
        index=True,
    )
    civora_next_action = fields.Char(
        string="Prochaine action",
        help="Prochaine action commerciale prevue (ex: rappeler, envoyer offre).",
    )

    # --- Coordonnees & localisation complementaires --------------------
    civora_whatsapp = fields.Char(string="WhatsApp")
    civora_neighborhood = fields.Char(
        string="Quartier",
        help="Quartier / zone (ex: Cocody, Marcory, Plateau).",
    )

    # --- Budget --------------------------------------------------------
    civora_currency_id = fields.Many2one(
        'res.currency',
        string="Devise",
        default=lambda self: self.env.company.currency_id,
    )
    civora_budget = fields.Monetary(
        string="Budget",
        currency_field='civora_currency_id',
        help="Budget estime du contact (achat / location).",
    )

    # --- Consentements RGPD --------------------------------------------
    civora_consent_email = fields.Selection(
        CIVORA_CONSENT, string="Consentement email", default='none',
    )
    civora_consent_sms = fields.Selection(
        CIVORA_CONSENT, string="Consentement SMS", default='none',
    )
    civora_consent_whatsapp = fields.Selection(
        CIVORA_CONSENT, string="Consentement WhatsApp", default='none',
    )

    # --- Interactions (timeline 360°) ----------------------------------
    civora_interaction_ids = fields.One2many(
        'civora.interaction', 'contact_id',
        string="Interactions",
    )
    civora_interaction_count = fields.Integer(
        string="Nb interactions",
        compute='_compute_civora_interaction_count',
    )
    civora_last_interaction_date = fields.Datetime(
        string="Dernière interaction",
        compute='_compute_civora_last_interaction_date',
        store=True,
        help="Date de la derniere interaction manuelle avec ce contact (hors evenements systeme).",
    )

    # --- Historique des roles ------------------------------------------
    civora_role_history_ids = fields.One2many(
        'civora.contact.role.history', 'partner_id',
        string="Historique des rôles",
    )

    # --- Journal des consentements RGPD --------------------------------
    civora_consent_log_ids = fields.One2many(
        'civora.consent.log', 'partner_id',
        string="Journal des consentements",
    )

    # --- Contraintes ---------------------------------------------------
    # Le score IA doit rester dans l'intervalle [0, 100].
    _civora_ai_score_range = models.Constraint(
        'CHECK(civora_ai_score >= 0 AND civora_ai_score <= 100)',
        "Le score IA doit etre compris entre 0 et 100.",
    )

    @api.depends('civora_role_ids', 'civora_role_ids.name')
    def _compute_civora_role_names(self):
        for partner in self:
            partner.civora_role_names = ", ".join(partner.civora_role_ids.mapped('name'))

    @api.depends('civora_interaction_ids', 'civora_interaction_ids.date', 'civora_interaction_ids.kind')
    def _compute_civora_last_interaction_date(self):
        """Dernière interaction MANUELLE (exclut les événements système
        comme changement de rôle)."""
        SYSTEM_KINDS = ('role_change',)
        for partner in self:
            manual = partner.civora_interaction_ids.filtered(
                lambda i: i.kind not in SYSTEM_KINDS
            )
            partner.civora_last_interaction_date = (
                max(manual.mapped('date')) if manual else False
            )

    def write(self, vals):
        """Capture des transitions de rôles + score + consentements RGPD.

        À chaque modification :
        - civora_role_ids / civora_primary_role_id → historisé dans role.history
          + interaction miroir de type 'role_change' dans la timeline
        - civora_consent_email/sms/whatsapp → journalisé dans consent.log
        """
        CONSENT_FIELDS = ('civora_consent_email', 'civora_consent_sms', 'civora_consent_whatsapp')
        # Snapshot AVANT écriture
        snapshot = {}
        consent_snapshot = {}
        if 'civora_role_ids' in vals or 'civora_primary_role_id' in vals:
            for p in self:
                snapshot[p.id] = {
                    'role_ids': set(p.civora_role_ids.ids),
                    'primary_id': p.civora_primary_role_id.id,
                }
        if any(f in vals for f in CONSENT_FIELDS):
            for p in self:
                consent_snapshot[p.id] = {
                    'civora_consent_email': p.civora_consent_email,
                    'civora_consent_sms': p.civora_consent_sms,
                    'civora_consent_whatsapp': p.civora_consent_whatsapp,
                }
        res = super().write(vals)
        # Après écriture, comparer et logger
        if snapshot:
            self._civora_log_role_changes(snapshot)
        if consent_snapshot:
            self._civora_log_consent_changes(consent_snapshot)
        return res

    def _civora_log_consent_changes(self, snapshot):
        """Enregistre les diffs de consentements dans consent.log."""
        Log = self.env['civora.consent.log'].sudo()
        CONSENT_CHANNELS = {
            'civora_consent_email': 'email',
            'civora_consent_sms': 'sms',
            'civora_consent_whatsapp': 'whatsapp',
        }
        # Source par défaut : manual (l'agent modifie via l'UI).
        # Si le contexte a un flag civora_consent_source, on l'utilise.
        source = self.env.context.get('civora_consent_source', 'manual')
        for partner in self:
            snap = snapshot.get(partner.id)
            if not snap:
                continue
            for field, channel in CONSENT_CHANNELS.items():
                old = snap[field]
                new = getattr(partner, field)
                if old != new:
                    Log.create({
                        'partner_id': partner.id,
                        'channel': channel,
                        'old_value': old or 'none',
                        'new_value': new or 'none',
                        'date': fields.Datetime.now(),
                        'user_id': self.env.user.id,
                        'source': source,
                    })

    def _civora_log_role_changes(self, snapshot):
        """Enregistre les diffs de rôles dans role.history + timeline."""
        History = self.env['civora.contact.role.history'].sudo()
        Interaction = self.env['civora.interaction'].sudo()
        Role = self.env['civora.contact.role']
        for partner in self:
            snap = snapshot.get(partner.id)
            if not snap:
                continue
            old_roles = snap['role_ids']
            new_roles = set(partner.civora_role_ids.ids)
            added = new_roles - old_roles
            removed = old_roles - new_roles
            old_primary = snap['primary_id']
            new_primary = partner.civora_primary_role_id.id

            events = []
            for rid in added:
                role = Role.browse(rid)
                History.create({
                    'partner_id': partner.id,
                    'role_id': rid,
                    'action': 'added',
                    'user_id': self.env.user.id,
                })
                events.append(('added', role.name))
            for rid in removed:
                role = Role.browse(rid)
                History.create({
                    'partner_id': partner.id,
                    'role_id': rid,
                    'action': 'removed',
                    'user_id': self.env.user.id,
                })
                events.append(('removed', role.name))
            if new_primary and new_primary != old_primary:
                role = Role.browse(new_primary)
                History.create({
                    'partner_id': partner.id,
                    'role_id': new_primary,
                    'action': 'set_primary',
                    'user_id': self.env.user.id,
                })
                events.append(('set_primary', role.name))

            # Créer 1 interaction miroir groupée (plus lisible que N entrées)
            if events:
                labels = {
                    'added': "Rôle ajouté",
                    'removed': "Rôle retiré",
                    'set_primary': "Rôle principal défini",
                }
                title_parts = [f"{labels[a]} : {n}" for a, n in events]
                Interaction.create({
                    'contact_id': partner.id,
                    'kind': 'role_change',
                    'title': " · ".join(title_parts),
                    'description': "",
                    'agent_id': self.env.user.id,
                    'date': fields.Datetime.now(),
                    'company_id': partner.company_id.id or self.env.company.id,
                })

    # ══════════════════════════════════════════════════════════════════
    # Score IA — calcul automatique
    # ══════════════════════════════════════════════════════════════════
    def civora_compute_ai_score(self):
        """Calcule le score IA (0-100) pour les contacts CIVORA du recordset.

        Formule pondérée sur 5 composants :
        - Activité récente     : 30 pts
        - Engagement financier : 25 pts
        - Complétude données   : 15 pts
        - Consentements RGPD   : 10 pts
        - Ancienneté / fidélité: 20 pts

        Bonus / malus selon statut CRM.
        Respecte le verrou manuel (civora_ai_score_manual = True → skip).
        """
        import json
        from datetime import timedelta
        now = fields.Datetime.now()
        today = fields.Date.context_today(self)
        Property = self.env.get('civora.property')
        Lease = self.env.get('civora.lease')

        for p in self:
            if not p.civora_is_contact and not p.civora_role_ids:
                continue
            if p.civora_ai_score_manual:
                continue

            # --- 1) Activité récente (max 30) ---
            interactions = p.civora_interaction_ids.filtered(
                lambda i: i.kind != 'role_change'
            )
            r30 = r60 = r90 = 0
            for it in interactions:
                if not it.date:
                    continue
                delta = (now - it.date).days
                if delta <= 30:
                    r30 += 1
                elif delta <= 60:
                    r60 += 1
                elif delta <= 90:
                    r90 += 1
            activity = min(30, r30 * 10 + r60 * 6 + r90 * 3)

            # --- 2) Engagement financier (max 25) ---
            engagement = 0
            has_property = False
            has_active_lease = False
            if Property is not None:
                has_property = bool(Property.sudo().search_count(
                    [('owner_id', '=', p.id)], limit=1
                ))
            if has_property:
                engagement += 10
            if Lease is not None:
                has_active_lease = bool(Lease.sudo().search_count([
                    '|',
                    ('tenant_id', '=', p.id),
                    ('owner_id', '=', p.id),
                    ('status', 'in', ['actif', 'retard', 'expire_bientot']),
                ], limit=1))
            if has_active_lease:
                engagement += 10
            if p.civora_budget and p.civora_budget > 0:
                engagement += 5

            # --- 3) Complétude des données (max 15) ---
            completeness = 0
            if p.email: completeness += 3
            if p.phone: completeness += 3
            if p.civora_primary_role_id: completeness += 3
            if p.civora_source_id: completeness += 3
            if p.civora_agent_id: completeness += 3

            # --- 4) Consentements RGPD (max 10) ---
            consent = 0
            if p.civora_consent_email == 'opt_in': consent += 4
            if p.civora_consent_sms == 'opt_in': consent += 3
            if p.civora_consent_whatsapp == 'opt_in': consent += 3

            # --- 5) Ancienneté (max 20) ---
            seniority = 0
            if p.create_date:
                years = (today - p.create_date.date()).days / 365.25
                if years >= 3: seniority = 20
                elif years >= 2: seniority = 15
                elif years >= 1: seniority = 10
                else: seniority = int(years * 10)

            # Total brut
            total = activity + engagement + completeness + consent + seniority

            # Bonus / malus statut CRM
            status_adj = 0
            if p.civora_status == 'chaud':
                status_adj = +10
            elif p.civora_status == 'inactif':
                total = min(total, 40)
            elif p.civora_status == 'a_risque':
                total = min(total, 60)
                # Malus si impayés multiples détectés
                if Lease is not None:
                    late = Lease.sudo().search_count([
                        '|',
                        ('tenant_id', '=', p.id),
                        ('owner_id', '=', p.id),
                        ('status', '=', 'retard'),
                    ])
                    if late >= 2:
                        status_adj -= 15

            total = max(0, min(100, total + status_adj))

            breakdown = {
                'activity': activity,
                'engagement': engagement,
                'completeness': completeness,
                'consent': consent,
                'seniority': seniority,
                'status_adj': status_adj,
                'total': total,
                'details': {
                    'interactions_30d': r30,
                    'interactions_60d': r60,
                    'interactions_90d': r90,
                    'has_property': has_property,
                    'has_active_lease': has_active_lease,
                    'has_budget': bool(p.civora_budget),
                    'years_since_create': round((today - p.create_date.date()).days / 365.25, 1) if p.create_date else 0,
                    'status': p.civora_status or '',
                },
            }
            p.sudo().write({
                'civora_ai_score': total,
                'civora_ai_score_breakdown': json.dumps(breakdown, ensure_ascii=False),
                'civora_ai_score_updated': now,
            })
        return True

    # RPC : version bouton "Recalculer" depuis la fiche 360°
    @api.model
    def civora_recompute_score(self, contact_id):
        """Force le recalcul du score pour un contact (ignore le verrou)."""
        p = self.browse(int(contact_id))
        if not p.exists():
            return {'success': False, 'error': "Contact introuvable."}
        # Bypass temporaire du verrou pour ce recalcul explicite
        was_manual = p.civora_ai_score_manual
        if was_manual:
            p.sudo().civora_ai_score_manual = False
        try:
            p.civora_compute_ai_score()
        finally:
            if was_manual:
                p.sudo().civora_ai_score_manual = True
        return {
            'success': True,
            'score': p.civora_ai_score,
            'breakdown': p.civora_ai_score_breakdown or '',
            'updated': str(p.civora_ai_score_updated) if p.civora_ai_score_updated else '',
        }

    @api.model
    def civora_toggle_score_lock(self, contact_id, locked):
        """Verrouille ou déverrouille le score IA pour un contact."""
        p = self.browse(int(contact_id))
        if not p.exists():
            return {'success': False}
        p.sudo().civora_ai_score_manual = bool(locked)
        return {'success': True, 'locked': p.civora_ai_score_manual}

    # Cron : recalcul nocturne
    @api.model
    def cron_civora_recompute_all_scores(self):
        """Recalcule le score IA pour tous les contacts CIVORA non verrouillés.
        Batché par 500 pour éviter les timeouts sur les grandes bases."""
        contacts = self.search([
            ('civora_is_contact', '=', True),
            ('civora_ai_score_manual', '=', False),
        ])
        total = len(contacts)
        batch_size = 500
        for i in range(0, total, batch_size):
            batch = contacts[i:i + batch_size]
            batch.civora_compute_ai_score()
            self.env.cr.commit()
        return {'processed': total}

    # ══════════════════════════════════════════════════════════════════
    # v10.2.0 — RGPD & Consentements
    # ══════════════════════════════════════════════════════════════════
    @api.model
    def civora_get_rgpd_registry(self, offset=0, limit=100, filter_channel=None, filter_value=None):
        """Retourne le registre des consentements pour l'onglet RGPD.

        - KPIs globaux (opt-in totaux, demandes traitées, données à MAJ, conformité)
        - Liste paginée des contacts avec leurs 3 consentements et date du dernier changement
        """
        domain = [('civora_is_contact', '=', True)]
        if filter_channel and filter_value:
            field = 'civora_consent_' + filter_channel
            if field in self._fields:
                domain.append((field, '=', filter_value))

        total = self.search_count(domain)

        # KPIs (portée globale, indépendante du filtre)
        base_domain = [('civora_is_contact', '=', True)]
        opt_in_email = self.search_count(base_domain + [('civora_consent_email', '=', 'opt_in')])
        opt_in_sms = self.search_count(base_domain + [('civora_consent_sms', '=', 'opt_in')])
        opt_in_wa = self.search_count(base_domain + [('civora_consent_whatsapp', '=', 'opt_in')])
        total_contacts = self.search_count(base_domain)
        # Contacts avec au moins un opt-in marketing (email/sms/wa)
        opt_in_any = self.search_count(base_domain + [
            '|', '|',
            ('civora_consent_email', '=', 'opt_in'),
            ('civora_consent_sms', '=', 'opt_in'),
            ('civora_consent_whatsapp', '=', 'opt_in'),
        ])
        # Contacts à mettre à jour = tous les 3 en 'none'
        to_update = self.search_count(base_domain + [
            ('civora_consent_email', '=', 'none'),
            ('civora_consent_sms', '=', 'none'),
            ('civora_consent_whatsapp', '=', 'none'),
        ])
        # Conformité globale = 1 - (à MAJ / total)
        conformity = 100 if not total_contacts else round((1 - to_update / total_contacts) * 100)

        records = self.search(domain, offset=offset, limit=limit, order='write_date desc')

        # Dernier événement consentement pour chaque contact
        Log = self.env['civora.consent.log'].sudo()
        rows = []
        for r in records:
            last = Log.search(
                [('partner_id', '=', r.id)], order='date desc', limit=1
            )
            rows.append({
                'id': r.id,
                'name': r.display_name or '',
                'email': r.email or '',
                'phone': r.phone or '',
                'city': r.city or '',
                'consent_email': r.civora_consent_email or 'none',
                'consent_sms': r.civora_consent_sms or 'none',
                'consent_whatsapp': r.civora_consent_whatsapp or 'none',
                'is_anonymized': (r.name or '').startswith('*** ANONYMISÉ'),
                'last_consent_date': str(last.date) if last else '',
            })

        return {
            'kpis': {
                'opt_in_any': opt_in_any,
                'opt_in_any_pct': 0 if not total_contacts else round(opt_in_any / total_contacts * 100, 1),
                'opt_in_email': opt_in_email,
                'opt_in_sms': opt_in_sms,
                'opt_in_whatsapp': opt_in_wa,
                'to_update': to_update,
                'conformity': conformity,
                'total_contacts': total_contacts,
            },
            'rows': rows,
            'total': total,
            'offset': offset,
        }

    @api.model
    def civora_get_consent_history(self, contact_id):
        """Retourne l'historique complet des consentements d'un contact."""
        Log = self.env['civora.consent.log'].sudo()
        events = Log.search([('partner_id', '=', int(contact_id))], order='date desc')
        CHANNEL_LABELS = {'email': "Email", 'sms': "SMS", 'whatsapp': "WhatsApp"}
        VALUE_LABELS = {'opt_in': "Opt-in", 'opt_out': "Opt-out", 'none': "Non renseigné"}
        SOURCE_LABELS = {
            'manual': "Saisie manuelle", 'import': "Import",
            'portal': "Portail", 'api': "API", 'system': "Système",
        }
        return [{
            'id': e.id,
            'channel': e.channel,
            'channel_label': CHANNEL_LABELS.get(e.channel, e.channel),
            'old_value': e.old_value or 'none',
            'old_label': VALUE_LABELS.get(e.old_value or 'none', e.old_value or '—'),
            'new_value': e.new_value,
            'new_label': VALUE_LABELS.get(e.new_value, e.new_value),
            'date': str(e.date) if e.date else '',
            'user_name': e.user_id.name if e.user_id else '',
            'source': e.source,
            'source_label': SOURCE_LABELS.get(e.source, e.source),
            'note': e.note or '',
        } for e in events]

    @api.model
    def civora_export_dpo(self, domain=None):
        """Export CSV DPO complet : contacts + consentements actifs + date du
        dernier changement + source. Format compatible régistre RGPD.
        """
        import base64, csv, io
        from datetime import date
        Log = self.env['civora.consent.log'].sudo()

        domain = domain or [('civora_is_contact', '=', True)]
        records = self.search(domain, order='name asc')

        headers = [
            'ID', 'Nom', 'Email', 'Téléphone', 'Ville',
            'Consent. Email', 'Consent. SMS', 'Consent. WhatsApp',
            'Dernier changement Email', 'Source Email',
            'Dernier changement SMS', 'Source SMS',
            'Dernier changement WhatsApp', 'Source WhatsApp',
            'Anonymisé', 'Créé le',
        ]
        consent_labels = {'opt_in': 'Opt-in', 'opt_out': 'Opt-out', 'none': 'Non renseigné'}
        source_labels = {
            'manual': "Saisie manuelle", 'import': "Import",
            'portal': "Portail", 'api': "API", 'system': "Système",
        }

        buf = io.StringIO()
        buf.write('\ufeff')  # BOM UTF-8 pour Excel
        writer = csv.writer(buf, delimiter=';', quoting=csv.QUOTE_MINIMAL)
        writer.writerow(headers)
        for r in records:
            # Récupérer le dernier changement pour chaque canal
            last_by_channel = {}
            for channel in ('email', 'sms', 'whatsapp'):
                last = Log.search([
                    ('partner_id', '=', r.id),
                    ('channel', '=', channel),
                ], order='date desc', limit=1)
                if last:
                    last_by_channel[channel] = (str(last.date.date()), source_labels.get(last.source, last.source))
                else:
                    last_by_channel[channel] = ('', '')
            writer.writerow([
                r.id,
                r.name or '',
                r.email or '',
                r.phone or '',
                r.city or '',
                consent_labels.get(r.civora_consent_email, '—'),
                consent_labels.get(r.civora_consent_sms, '—'),
                consent_labels.get(r.civora_consent_whatsapp, '—'),
                last_by_channel['email'][0], last_by_channel['email'][1],
                last_by_channel['sms'][0], last_by_channel['sms'][1],
                last_by_channel['whatsapp'][0], last_by_channel['whatsapp'][1],
                'Oui' if (r.name or '').startswith('*** ANONYMISÉ') else 'Non',
                str(r.create_date.date()) if r.create_date else '',
            ])

        raw = buf.getvalue().encode('utf-8')
        return {
            'content': base64.b64encode(raw).decode('ascii'),
            'filename': 'registre_rgpd_%s.csv' % date.today().strftime('%Y%m%d'),
            'count': len(records),
        }

    @api.model
    def civora_anonymize_contact(self, contact_id, reason=None):
        """Droit à l'oubli : anonymise le contact tout en préservant l'ID.

        - Nom : "*** ANONYMISÉ ***"
        - Email, phone, mobile, whatsapp, adresse : effacés
        - Comment : "Contact anonymisé le YYYY-MM-DD suite à demande RGPD"
        - civora_is_contact reste True (pour maintenir le cloisonnement)
        - Les rôles, statut, score sont conservés (statistiques agrégées)
        - Toutes les interactions sont conservées mais leur description est effacée
        - Un événement 'system' est logué dans le journal des consentements
        """
        from datetime import date
        p = self.browse(int(contact_id))
        if not p.exists():
            return {'success': False, 'error': "Contact introuvable."}
        if (p.name or '').startswith('*** ANONYMISÉ'):
            return {'success': False, 'error': "Ce contact est déjà anonymisé."}

        today_str = date.today().strftime('%Y-%m-%d')
        # Détruire les interactions verbeuses (contenu personnel) mais garder les lignes
        for it in p.civora_interaction_ids:
            it.sudo().write({
                'title': "[Contenu anonymisé]",
                'description': "",
            })

        # Anonymiser le contact lui-même
        p.sudo().write({
            'name': '*** ANONYMISÉ (%s) ***' % today_str,
            'email': False,
            'phone': False,
            'civora_whatsapp': False,
            'street': False,
            'street2': False,
            'zip': False,
            'civora_neighborhood': False,
            'company_name': False,
            'comment': "Contact anonymisé le %s suite à demande RGPD. Raison : %s"
                       % (today_str, reason or 'Non précisée'),
            'civora_consent_email': 'opt_out',
            'civora_consent_sms': 'opt_out',
            'civora_consent_whatsapp': 'opt_out',
        })

        # Journaliser
        self.env['civora.consent.log'].sudo().create({
            'partner_id': p.id,
            'channel': 'email',
            'old_value': 'opt_in',
            'new_value': 'opt_out',
            'date': fields.Datetime.now(),
            'user_id': self.env.user.id,
            'source': 'system',
            'note': "Anonymisation (droit à l'oubli) : %s" % (reason or 'Non précisée'),
        })
        return {'success': True, 'contact_id': p.id}

    @api.model
    def civora_export_personal_data(self, contact_id):
        """Export RGPD des données d'un contact au format JSON.

        Contient toutes les données personnelles connues + interactions +
        historique rôles + consentements. Format destiné à répondre au droit
        d'accès RGPD (article 15).
        """
        import base64, json
        from datetime import date
        p = self.browse(int(contact_id))
        if not p.exists():
            return {'success': False, 'error': "Contact introuvable."}

        Log = self.env['civora.consent.log'].sudo()
        History = self.env['civora.contact.role.history'].sudo()

        data = {
            'export_date': str(fields.Datetime.now()),
            'contact': {
                'id': p.id,
                'name': p.name or '',
                'email': p.email or '',
                'phone': p.phone or '',
                'whatsapp': p.civora_whatsapp or '',
                'city': p.city or '',
                'neighborhood': p.civora_neighborhood or '',
                'street': p.street or '',
                'company': p.company_name or '',
                'roles': p.civora_role_ids.mapped('name'),
                'primary_role': p.civora_primary_role_id.name if p.civora_primary_role_id else '',
                'source': p.civora_source_id.name if p.civora_source_id else '',
                'status': p.civora_status or '',
                'ai_score': p.civora_ai_score or 0,
                'budget': p.civora_budget or 0,
                'agent': p.civora_agent_id.name if p.civora_agent_id else '',
                'created': str(p.create_date) if p.create_date else '',
                'consents': {
                    'email': p.civora_consent_email or 'none',
                    'sms': p.civora_consent_sms or 'none',
                    'whatsapp': p.civora_consent_whatsapp or 'none',
                },
            },
            'interactions': [{
                'id': i.id, 'kind': i.kind, 'title': i.title or '',
                'description': i.description or '',
                'date': str(i.date) if i.date else '',
                'agent': i.agent_id.name if i.agent_id else '',
            } for i in p.civora_interaction_ids],
            'role_history': [{
                'role': h.role_id.name if h.role_id else '',
                'action': h.action,
                'date': str(h.date) if h.date else '',
                'user': h.user_id.name if h.user_id else '',
                'note': h.note or '',
            } for h in History.search([('partner_id', '=', p.id)])],
            'consent_history': [{
                'channel': l.channel,
                'old': l.old_value or 'none',
                'new': l.new_value,
                'date': str(l.date) if l.date else '',
                'user': l.user_id.name if l.user_id else '',
                'source': l.source,
                'note': l.note or '',
            } for l in Log.search([('partner_id', '=', p.id)])],
        }
        raw = json.dumps(data, ensure_ascii=False, indent=2).encode('utf-8')
        return {
            'success': True,
            'content': base64.b64encode(raw).decode('ascii'),
            'filename': 'donnees_perso_%d_%s.json' % (p.id, date.today().strftime('%Y%m%d')),
        }

    # ══════════════════════════════════════════════════════════════════
    # v10.2.0 — Imports CSV
    # ══════════════════════════════════════════════════════════════════
    @api.model
    def civora_import_preview(self, file_base64, filename):
        """Parse le fichier CSV/Excel et retourne :
        - Les headers détectés
        - Les 5 premières lignes (preview)
        - Le mapping automatique proposé (header → champ_civora)
        Ne crée AUCUN contact — juste un aperçu.
        """
        import base64, csv, io
        FIELD_ALIASES = {
            'name': ['nom', 'name', 'contact', 'nom complet', 'full name', 'raison sociale'],
            'email': ['email', 'e-mail', 'mail', 'courriel'],
            'phone': ['telephone', 'téléphone', 'phone', 'tel', 'mobile', 'gsm', 'contact', 'numero'],
            'civora_whatsapp': ['whatsapp', 'wa', 'whats'],
            'city': ['ville', 'city'],
            'civora_neighborhood': ['quartier', 'zone', 'commune'],
            'street': ['adresse', 'address', 'rue'],
            'company_name': ['societe', 'société', 'company', 'entreprise', 'compagnie'],
        }
        try:
            raw = base64.b64decode(file_base64)
        except Exception as e:
            return {'success': False, 'error': "Fichier illisible : %s" % str(e)[:100]}

        # Détection format
        is_excel = filename.lower().endswith(('.xlsx', '.xls'))
        rows = []
        headers = []
        try:
            if is_excel:
                try:
                    import openpyxl
                except ImportError:
                    return {'success': False, 'error': "Module openpyxl requis pour lire les fichiers Excel. Convertissez en CSV."}
                wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
                ws = wb.active
                iter_rows = list(ws.iter_rows(values_only=True))
                if not iter_rows:
                    return {'success': False, 'error': "Fichier vide."}
                headers = [str(c or '').strip() for c in iter_rows[0]]
                for row in iter_rows[1:]:
                    rows.append([str(c or '').strip() if c is not None else '' for c in row])
            else:
                # CSV — détection auto du séparateur
                text = raw.decode('utf-8-sig', errors='replace')
                sniffer = csv.Sniffer()
                sample = text[:1024]
                try:
                    dialect = sniffer.sniff(sample, delimiters=';,\t|')
                    sep = dialect.delimiter
                except Exception:
                    sep = ';' if text.count(';') > text.count(',') else ','
                reader = csv.reader(io.StringIO(text), delimiter=sep)
                iter_rows = list(reader)
                if not iter_rows:
                    return {'success': False, 'error': "Fichier vide."}
                headers = [h.strip() for h in iter_rows[0]]
                rows = [list(r) for r in iter_rows[1:] if any(c.strip() for c in r)]
        except Exception as e:
            return {'success': False, 'error': "Erreur de lecture : %s" % str(e)[:200]}

        # Auto-mapping
        mapping = {}
        for i, h in enumerate(headers):
            h_norm = h.lower().strip()
            for field, aliases in FIELD_ALIASES.items():
                if h_norm in aliases or any(a in h_norm for a in aliases):
                    if field not in mapping.values():  # premier arrivé, premier servi
                        mapping[str(i)] = field
                        break

        preview = rows[:5]
        return {
            'success': True,
            'headers': headers,
            'mapping': mapping,
            'preview': preview,
            'total_rows': len(rows),
            'available_fields': [
                {'code': 'name', 'label': 'Nom (obligatoire)'},
                {'code': 'email', 'label': 'Email'},
                {'code': 'phone', 'label': 'Téléphone'},
                {'code': 'civora_whatsapp', 'label': 'WhatsApp'},
                {'code': 'city', 'label': 'Ville'},
                {'code': 'civora_neighborhood', 'label': 'Quartier'},
                {'code': 'street', 'label': 'Adresse'},
                {'code': 'company_name', 'label': 'Société'},
            ],
        }

    @api.model
    def civora_import_run(self, file_base64, filename, mapping, skip_duplicates=True):
        """Exécute l'import selon le mapping validé par l'utilisateur.

        mapping : {"0": "name", "1": "email", ...} (index colonne → champ CIVORA)
        skip_duplicates : si True, un doublon détecté est ignoré. Sinon, créé quand même.

        Retourne : {imported, duplicates, errors, import_id}
        """
        import base64, csv, io, json
        try:
            raw = base64.b64decode(file_base64)
        except Exception as e:
            return {'success': False, 'error': "Fichier illisible."}
        is_excel = filename.lower().endswith(('.xlsx', '.xls'))
        rows = []
        try:
            if is_excel:
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
                ws = wb.active
                iter_rows = list(ws.iter_rows(values_only=True))
                for row in iter_rows[1:]:
                    rows.append([str(c or '').strip() if c is not None else '' for c in row])
            else:
                text = raw.decode('utf-8-sig', errors='replace')
                try:
                    dialect = csv.Sniffer().sniff(text[:1024], delimiters=';,\t|')
                    sep = dialect.delimiter
                except Exception:
                    sep = ';' if text.count(';') > text.count(',') else ','
                reader = csv.reader(io.StringIO(text), delimiter=sep)
                iter_rows = list(reader)
                rows = [list(r) for r in iter_rows[1:] if any(c.strip() for c in r)]
        except Exception as e:
            return {'success': False, 'error': "Erreur de lecture : %s" % str(e)[:200]}

        # Créer l'enregistrement d'import
        import_rec = self.env['civora.contact.import'].sudo().create({
            'name': filename,
            'file_name': filename,
            'file_data': file_base64,
            'mapping_json': json.dumps(mapping),
            'state': 'running',
            'total_rows': len(rows),
        })

        # Trouver quel champ est mappé au nom (obligatoire)
        name_col = None
        for col_idx, field in mapping.items():
            if field == 'name':
                name_col = int(col_idx)
                break
        if name_col is None:
            import_rec.sudo().write({'state': 'error', 'error_log': "Aucune colonne 'Nom' mappée."})
            return {'success': False, 'error': "Vous devez mapper au moins la colonne Nom."}

        imported_ids = []
        duplicates = 0
        errors = 0
        errors_log = []

        for row_idx, row in enumerate(rows):
            try:
                # Construire vals depuis mapping
                vals = {'civora_is_contact': True}
                for col_idx_str, field in mapping.items():
                    col_idx = int(col_idx_str)
                    if col_idx >= len(row):
                        continue
                    v = row[col_idx].strip() if isinstance(row[col_idx], str) else row[col_idx]
                    if v:
                        vals[field] = v

                if not vals.get('name'):
                    errors += 1
                    errors_log.append("Ligne %d : nom vide" % (row_idx + 2))
                    continue

                # Dédoublonnage
                if skip_duplicates:
                    dups = self.civora_find_duplicates(
                        name=vals.get('name'),
                        email=vals.get('email'),
                        phone=vals.get('phone'),
                    )
                    if dups:
                        duplicates += 1
                        continue

                # Créer avec source 'import' pour les consentements
                created = self.with_context(civora_consent_source='import').create(vals)
                imported_ids.append(created.id)
            except Exception as e:
                errors += 1
                errors_log.append("Ligne %d : %s" % (row_idx + 2, str(e)[:100]))

        import_rec.sudo().write({
            'state': 'done',
            'imported_count': len(imported_ids),
            'duplicates_count': duplicates,
            'errors_count': errors,
            'error_log': '\n'.join(errors_log[:100]) if errors_log else '',
            'created_partner_ids': json.dumps(imported_ids),
        })

        return {
            'success': True,
            'imported': len(imported_ids),
            'duplicates': duplicates,
            'errors': errors,
            'total': len(rows),
            'import_id': import_rec.id,
        }

    @api.model
    def civora_get_import_history(self, limit=10):
        """Retourne les N derniers imports pour affichage dans l'onglet Imports."""
        Import = self.env['civora.contact.import'].sudo()
        recs = Import.search([], limit=limit, order='create_date desc')
        return [{
            'id': r.id,
            'name': r.name or '',
            'state': r.state,
            'total_rows': r.total_rows,
            'imported_count': r.imported_count,
            'duplicates_count': r.duplicates_count,
            'errors_count': r.errors_count,
            'user_name': r.user_id.name if r.user_id else '',
            'date': str(r.create_date) if r.create_date else '',
        } for r in recs]

    @api.model
    def civora_import_template(self):
        """Retourne un fichier CSV modèle prêt à remplir.

        - Colonnes officielles CIVORA (8) avec noms reconnus par l'auto-mapping
        - 2 lignes d'exemple pour montrer le format attendu (téléphones Côte d'Ivoire,
          quartiers d'Abidjan, etc.)
        - Séparateur ; (compatible Excel FR/CI)
        - UTF-8 BOM (accents lisibles dans Excel)

        L'utilisateur peut télécharger, remplir et déposer dans la dropzone —
        l'auto-mapping reconnaîtra 100% des colonnes.
        """
        import base64, csv, io
        from datetime import date

        headers = [
            "Nom",              # → name (obligatoire)
            "Email",            # → email
            "Téléphone",        # → phone
            "WhatsApp",         # → civora_whatsapp
            "Société",          # → company_name
            "Ville",            # → city
            "Quartier",         # → civora_neighborhood
            "Adresse",          # → street
        ]
        examples = [
            [
                "Aké Brigitte",
                "brigitte.ake@example.ci",
                "+225 07 12 34 56 78",
                "+225 07 12 34 56 78",
                "",
                "Abidjan",
                "Cocody",
                "Riviera Golf, Rue des Jardins",
            ],
            [
                "Société Tech CI SARL",
                "contact@techci.com",
                "+225 27 22 44 55 66",
                "",
                "Tech CI SARL",
                "Abidjan",
                "Plateau",
                "Immeuble Sicogi, 3e étage",
            ],
        ]

        buf = io.StringIO()
        buf.write('\ufeff')  # BOM UTF-8 pour Excel
        writer = csv.writer(buf, delimiter=';', quoting=csv.QUOTE_MINIMAL)
        writer.writerow(headers)
        for row in examples:
            writer.writerow(row)

        raw = buf.getvalue().encode('utf-8')
        return {
            'content': base64.b64encode(raw).decode('ascii'),
            'filename': 'modele_import_contacts_civora_%s.csv' % date.today().strftime('%Y%m%d'),
        }

    @api.model_create_multi
    def create(self, vals_list):
        # Cloisonnement strict : tout contact CIVORA cree sans societe est
        # rattache a la societe courante. "Contact CIVORA" = flag civora_is_contact
        # OU au moins un role CIVORA (acquereur, proprietaire, locataire...).
        # Les partenaires non-CIVORA gardent le comportement natif d'Odoo.
        for vals in vals_list:
            if vals.get('company_id'):
                continue
            has_role = bool(vals.get('civora_role_ids'))
            if vals.get('civora_is_contact') or has_role:
                vals['company_id'] = self.env.company.id
        return super().create(vals_list)

    @api.depends('civora_interaction_ids')
    def _compute_civora_interaction_count(self):
        for partner in self:
            partner.civora_interaction_count = len(partner.civora_interaction_ids)

    @api.constrains('civora_primary_role_id', 'civora_role_ids')
    def _check_civora_primary_role(self):
        for partner in self:
            if (
                partner.civora_primary_role_id
                and partner.civora_primary_role_id not in partner.civora_role_ids
            ):
                raise ValidationError(_(
                    "Le role principal doit faire partie des roles du contact."
                ))

    # ══════════════════════════════════════════════════════════════════
    # RPC — Detection de doublons
    # ══════════════════════════════════════════════════════════════════
    @api.model
    def civora_find_duplicates(self, name=None, email=None, phone=None, exclude_id=None):
        """Retourne les contacts CIVORA potentiellement doublons.

        Critères (OR logique) :
        - Nom similaire (ilike) ET (email OU téléphone identique)
        - Email identique (si fourni, non vide)
        - Téléphone identique (si fourni, non vide)

        Utilisé par le drawer à la création pour prévenir l'utilisateur avant
        de saisir un contact déjà existant.
        """
        name = (name or '').strip()
        email = (email or '').strip().lower()
        phone = (phone or '').strip().replace(' ', '').replace('.', '').replace('-', '')

        sub_domains = []
        if email:
            sub_domains.append([('email', '=ilike', email)])
        if phone:
            # Match sur téléphone : on cherche les chiffres uniquement
            digits = ''.join(c for c in phone if c.isdigit())
            if len(digits) >= 6:
                sub_domains.append([('phone', 'ilike', digits[-8:])])
                sub_domains.append([('civora_whatsapp', 'ilike', digits[-8:])])
        if name and len(name) >= 3:
            sub_domains.append([('name', 'ilike', name)])

        if not sub_domains:
            return []

        # Combiner en OR polish
        base = [('civora_is_contact', '=', True)]
        if exclude_id:
            base.append(('id', '!=', int(exclude_id)))

        if len(sub_domains) == 1:
            domain = base + sub_domains[0]
        else:
            or_prefix = ['|'] * (len(sub_domains) - 1)
            # Chaque sub_domain à K leaves : préfixer par (K-1) &
            polish_subs = []
            for sd in sub_domains:
                k = len(sd)
                polish_subs.extend(['&'] * (k - 1))
                polish_subs.extend(sd)
            domain = base + or_prefix + polish_subs

        matches = self.search(domain, limit=10)
        return [{
            'id': m.id,
            'name': m.display_name or '',
            'email': m.email or '',
            'phone': m.phone or '',
            'city': m.city or '',
            'primary_role': m.civora_primary_role_id.name if m.civora_primary_role_id else '',
            'status': m.civora_status or '',
            'ai_score': m.civora_ai_score or 0,
        } for m in matches]

    # ══════════════════════════════════════════════════════════════════
    # RPC — Export CSV
    # ══════════════════════════════════════════════════════════════════
    @api.model
    def civora_export_csv(self, domain=None):
        """Exporte les contacts CIVORA d'un domaine donné au format CSV.

        Retourne { 'content': <base64>, 'filename': 'contacts_YYYYMMDD.csv' }
        Le client déclenche le téléchargement via une balise <a href data:...>.
        """
        import base64, csv, io
        from datetime import date

        domain = domain or [('civora_is_contact', '=', True)]
        records = self.search(domain, order='name asc')

        headers = [
            'Nom', 'Société', 'Email', 'Téléphone', 'WhatsApp',
            'Ville', 'Quartier', 'Adresse',
            'Rôle principal', 'Rôles', 'Source', 'Segments',
            'Statut', 'Score IA', 'Agent', 'Prochaine action', 'Budget',
            'Consentement Email', 'Consentement SMS', 'Consentement WhatsApp',
            'Nb interactions', 'Créé le', 'Modifié le',
        ]

        consent_labels = {'opt_in': 'Opt-in', 'opt_out': 'Opt-out', 'none': '—'}
        status_labels = {
            'chaud': 'Chaud', 'actif': 'Actif', 'qualifie': 'Qualifié',
            'a_risque': 'À risque', 'inactif': 'Inactif',
        }

        buf = io.StringIO()
        # UTF-8 BOM pour ouverture correcte dans Excel
        buf.write('\ufeff')
        writer = csv.writer(buf, delimiter=';', quoting=csv.QUOTE_MINIMAL)
        writer.writerow(headers)
        for r in records:
            writer.writerow([
                r.name or '',
                r.company_name or '',
                r.email or '',
                r.phone or '',
                r.civora_whatsapp or '',
                r.city or '',
                r.civora_neighborhood or '',
                r.street or '',
                r.civora_primary_role_id.name if r.civora_primary_role_id else '',
                r.civora_role_names or '',
                r.civora_source_id.name if r.civora_source_id else '',
                ', '.join(r.civora_segment_ids.mapped('name')) if r.civora_segment_ids else '',
                status_labels.get(r.civora_status, r.civora_status or ''),
                r.civora_ai_score or 0,
                r.civora_agent_id.name if r.civora_agent_id else '',
                r.civora_next_action or '',
                r.civora_budget or 0,
                consent_labels.get(r.civora_consent_email, '—'),
                consent_labels.get(r.civora_consent_sms, '—'),
                consent_labels.get(r.civora_consent_whatsapp, '—'),
                r.civora_interaction_count or 0,
                str(r.create_date.date()) if r.create_date else '',
                str(r.write_date.date()) if r.write_date else '',
            ])
        raw = buf.getvalue().encode('utf-8')
        return {
            'content': base64.b64encode(raw).decode('ascii'),
            'filename': 'contacts_civora_%s.csv' % date.today().strftime('%Y%m%d'),
            'count': len(records),
        }

    # ══════════════════════════════════════════════════════════════════
    # RPC — Liste des IDs pour navigation (siblings)
    # ══════════════════════════════════════════════════════════════════
    @api.model
    def civora_get_sibling_ids(self, domain=None, order='civora_ai_score desc, name asc'):
        """Retourne uniquement les IDs matchant le domaine, ordonnés.

        Utilisé par la fiche 360° pour la navigation précédent/suivant :
        l'annuaire passe le domaine actif et la fiche récupère la liste
        complète des IDs (juste les IDs, léger, ~40Ko pour 5000 contacts).
        """
        domain = domain or [('civora_is_contact', '=', True)]
        records = self.search(domain, order=order)
        return records.ids
